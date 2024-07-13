import openpyxl, os
from openpyxl.utils import get_column_letter


os.chdir(os.path.dirname(os.path.abspath(__file__)))


txtToExcelWB = openpyxl.Workbook()

sheet = txtToExcelWB.active


for folder, subfolders, txtFiles in os.walk('txtFiles'):
    for index, txtFile in enumerate(txtFiles, start = 1):
        oTxtFile = open(os.path.join(folder, txtFile))
        txtList = oTxtFile.readlines()
        
        for i, line in enumerate(txtList, start = 1): # By default it started from 0. Therefore we need to start from 1 since it's a spreadsheet.
            sheet[get_column_letter(index) + str(i)] = line
        
    

txtToExcelWB.save('resultSpreadsheetFromTxtFiles.xlsx')